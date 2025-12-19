import pandas as pd
import streamlit as st
from config import GRADE_ORDER, is_passing_grade, get_allowed_assignment_types

def read_progress_report(filepath):
    """
    Reads an uploaded progress report (Excel or CSV), in either:
      - “long” format with columns [ID or STUDENT ID, NAME, Course, Grade, Year, Semester]
      - “wide” format with columns ID/NAME plus COURSE_* or COURSE * columns
    Returns a long‐format DataFrame with exactly ['ID','NAME','Course','Grade','Year','Semester'] or None on error.
    """
    try:
        # Excel files
        if filepath.lower().endswith(('.xlsx', '.xls')):
            xls = pd.ExcelFile(filepath)
            # If there is a sheet literally named "Progress Report", read that long‐form
            if 'Progress Report' in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name='Progress Report')
                required = {'ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester'}
                missing = required - set(df.columns)
                if missing:
                    available_cols = ', '.join(sorted(df.columns))
                    missing_cols = ', '.join(sorted(missing))
                    st.error(
                        f"**Missing required columns:** {missing_cols}\n\n"
                        f"**Available columns:** {available_cols}\n\n"
                        f"**Expected format:** Long format with columns: ID, NAME, Course, Grade, Year, Semester\n"
                        f"**Tip:** Make sure your Excel sheet is named 'Progress Report' or contains these exact column names."
                    )
                    return None
                return df[list(required)]
            # Otherwise, pull the first sheet and check if it's already in long format
            df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
            # Check if it's already in long format (has Course, Grade, Year, Semester columns)
            if {'Course', 'Grade', 'Year', 'Semester'}.issubset(df.columns):
                # It's already in long format, just extract the needed columns
                # Handle ID column variations
                id_col = None
                if 'ID' in df.columns:
                    id_col = 'ID'
                elif 'STUDENT ID' in df.columns:
                    id_col = 'STUDENT ID'
                else:
                    available_cols = ', '.join(sorted(df.columns))
                    st.error(
                        f"**Missing ID column (Excel file)**\n\n"
                        f"**Available columns:** {available_cols}\n\n"
                        f"**Expected:** A column named 'ID' or 'STUDENT ID'\n"
                        f"**Tip:** Check if your ID column has a different name (e.g., 'Student ID', 'StudentID'). "
                        f"You may need to rename it in Excel before uploading."
                    )
                    return None
                # Handle NAME column variations
                name_col = None
                if 'NAME' in df.columns:
                    name_col = 'NAME'
                elif 'Name' in df.columns:
                    name_col = 'Name'
                else:
                    available_cols = ', '.join(sorted(df.columns))
                    st.error(
                        f"**Missing NAME column (Excel file)**\n\n"
                        f"**Available columns:** {available_cols}\n\n"
                        f"**Expected:** A column named 'NAME' or 'Name'\n"
                        f"**Tip:** Check if your name column has a different name (e.g., 'Student Name', 'Full Name'). "
                        f"You may need to rename it in Excel before uploading."
                    )
                    return None
                # Select and rename columns to standard format
                result_df = df[[id_col, name_col, 'Course', 'Grade', 'Year', 'Semester']].copy()
                result_df = result_df.rename(columns={id_col: 'ID', name_col: 'NAME'})
                return result_df
            # Otherwise, attempt to transform wide → long
            # First, detect format and provide feedback
            detected_format = _detect_file_format(df)
            if detected_format:
                st.info(f"📋 **Detected format:** {detected_format}")
            
            transformed = transform_wide_format(df)
            if transformed is None:
                available_cols = ', '.join(sorted(df.columns))
                st.error(
                    f"**Failed to read the uploaded progress report file.**\n\n"
                    f"**Available columns:** {available_cols}\n\n"
                    f"**Possible issues:**\n"
                    f"- File is not in the expected format (long or wide)\n"
                    f"- Wide format files should have COURSE_* columns with values like 'CODE/SEM-YYYY/GRADE'\n"
                    f"- Long format files should have columns: ID, NAME, Course, Grade, Year, Semester\n"
                    f"**Tip:** If your file is in long format, ensure it has a sheet named 'Progress Report' or rename your columns to match the expected names."
                )
            return transformed

        # CSV files
        elif filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
            if {'Course','Grade','Year','Semester'}.issubset(df.columns):
                # It's already in long format, just extract the needed columns
                # Handle ID column variations
                id_col = None
                if 'ID' in df.columns:
                    id_col = 'ID'
                elif 'STUDENT ID' in df.columns:
                    id_col = 'STUDENT ID'
                else:
                    available_cols = ', '.join(sorted(df.columns))
                    st.error(
                        f"**Missing ID column (CSV file)**\n\n"
                        f"**Available columns:** {available_cols}\n\n"
                        f"**Expected:** A column named 'ID' or 'STUDENT ID'\n"
                        f"**Tip:** Check if your ID column has a different name. "
                        f"You may need to rename it before uploading."
                    )
                    return None
                # Handle NAME column variations
                name_col = None
                if 'NAME' in df.columns:
                    name_col = 'NAME'
                elif 'Name' in df.columns:
                    name_col = 'Name'
                else:
                    available_cols = ', '.join(sorted(df.columns))
                    st.error(
                        f"**Missing NAME column (CSV file)**\n\n"
                        f"**Available columns:** {available_cols}\n\n"
                        f"**Expected:** A column named 'NAME' or 'Name'\n"
                        f"**Tip:** Check if your name column has a different name. "
                        f"You may need to rename it before uploading."
                    )
                    return None
                # Select and rename columns to standard format
                result_df = df[[id_col, name_col, 'Course', 'Grade', 'Year', 'Semester']].copy()
                result_df = result_df.rename(columns={id_col: 'ID', name_col: 'NAME'})
                return result_df
            # otherwise try wide form
            # First, detect format and provide feedback
            detected_format = _detect_file_format(df)
            if detected_format:
                st.info(f"📋 **Detected format:** {detected_format}")
            
            transformed = transform_wide_format(df)
            if transformed is None:
                available_cols = ', '.join(sorted(df.columns))
                st.error(
                    f"**Failed to read the uploaded progress report file.**\n\n"
                    f"**Available columns:** {available_cols}\n\n"
                    f"**Possible issues:**\n"
                    f"- File is not in the expected format (long or wide)\n"
                    f"- Wide format files should have COURSE_* columns with values like 'CODE/SEM-YYYY/GRADE'\n"
                    f"- Long format files should have columns: ID, NAME, Course, Grade, Year, Semester\n"
                    f"**Tip:** Ensure your CSV has the correct column names or format."
                )
            return transformed

        else:
            st.error("Unsupported file format. Upload an Excel or CSV.")
            return None

    except Exception as e:
        error_msg = str(e)
        st.error(
            f"**Error reading file:** {error_msg}\n\n"
            f"**Common causes:**\n"
            f"- File is corrupted or password-protected\n"
            f"- File format is not supported (only .xlsx, .xls, .csv are supported)\n"
            f"- File is too large or has formatting issues\n"
            f"**Tip:** Try opening the file in Excel to verify it's not corrupted, then save it again."
        )
        return None


def _detect_file_format(df: pd.DataFrame) -> str:
    """
    Detects the file format and returns a descriptive string.
    """
    if {'Course', 'Grade', 'Year', 'Semester'}.issubset(df.columns):
        return "Long format (detected columns: Course, Grade, Year, Semester)"
    elif any(c.upper().startswith('COURSE') for c in df.columns):
        return "Wide format (detected COURSE_* columns)"
    else:
        return "Unknown format (does not match expected long or wide format)"


def transform_wide_format(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Converts a wide‐format progress sheet into long form.
    Detects an ID column ('ID' or 'STUDENT ID') plus any 'COURSE…' columns.
    Expects each cell as 'COURSECODE/SEMESTER-YEAR/GRADE'.
    Returns a DataFrame with columns ['ID','NAME','Course','Grade','Year','Semester'].
    """
    # 1) Find the student‐ID column
    if 'ID' in df.columns:
        id_col = 'ID'
    elif 'STUDENT ID' in df.columns:
        id_col = 'STUDENT ID'
    else:
        st.error("Wide format file missing an 'ID' or 'STUDENT ID' column.")
        return None

    # 2) Ensure we have a NAME column
    if 'NAME' not in df.columns and 'Name' in df.columns:
        df = df.rename(columns={'Name': 'NAME'})
    if 'NAME' not in df.columns:
        st.error("Wide format file missing a 'NAME' column.")
        return None

    # 3) Detect all COURSE columns
    course_cols = [c for c in df.columns if c.upper().startswith('COURSE')]
    if not course_cols:
        st.error("Wide format file missing any 'COURSE…' columns.")
        return None

    # 4) Melt into long form
    id_vars = [c for c in df.columns if c not in course_cols]
    df_melt = df.melt(
        id_vars=id_vars,
        value_vars=course_cols,
        var_name='Course_Column',
        value_name='CourseData'
    )

    # 5) Drop empty entries
    df_melt = df_melt[df_melt['CourseData'].notna() & (df_melt['CourseData'].str.strip() != '')]

    # 6) Split COURSECODE/SEMESTER-YEAR/GRADE
    parts = df_melt['CourseData'].str.split('/', expand=True)
    if parts.shape[1] < 3:
        # Show sample of problematic data
        sample_data = df_melt['CourseData'].dropna().head(3).tolist()
        sample_str = '\n'.join([f"- {val}" for val in sample_data])
        st.error(
            f"**Parsing error: Expected format 'CODE/SEM-YYYY/GRADE'**\n\n"
            f"**Sample data found:**\n{sample_str}\n\n"
            f"**Expected format:** Each course cell should contain: COURSECODE/SEMESTER-YEAR/GRADE\n"
            f"**Example:** CHEM201/Fall-2022/B+\n"
            f"**Tip:** Check that your course columns contain data in the format: CourseCode/Semester-Year/Grade"
        )
        return None

    df_melt['Course'] = parts[0].str.strip().str.upper()
    df_melt['RawSemYear'] = parts[1].str.strip()
    df_melt['Grade'] = parts[2].str.strip().str.upper()

    # 7) Split Semester and Year
    sem_parts = df_melt['RawSemYear'].str.split('-', expand=True)
    if sem_parts.shape[1] < 2:
        # Show sample of problematic data
        sample_data = df_melt['RawSemYear'].dropna().head(3).tolist()
        sample_str = '\n'.join([f"- {val}" for val in sample_data])
        st.error(
            f"**Parsing error: Expected Semester-Year format 'SEMESTER-YYYY'**\n\n"
            f"**Sample data found:**\n{sample_str}\n\n"
            f"**Expected format:** Semester-Year should be: SEMESTER-YYYY\n"
            f"**Examples:** Fall-2022, Spring-2023, Summer-2024\n"
            f"**Tip:** Ensure the semester and year are separated by a hyphen (e.g., Fall-2022, not Fall 2022)"
        )
        return None

    df_melt['Semester'] = sem_parts[0].str.strip().str.title()
    df_melt['Year'] = sem_parts[1].str.strip()

    # 8) Rename ID column to 'ID'
    if id_col != 'ID':
        df_melt = df_melt.rename(columns={id_col: 'ID'})

    # 9) Collect only needed columns
    final_cols = ['ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester']
    missing = [c for c in final_cols if c not in df_melt.columns]
    if missing:
        st.error(f"Missing columns after transformation: {missing}")
        return None

    return df_melt[final_cols].drop_duplicates()


def read_equivalent_courses(equivalent_courses_df):
    mapping = {}
    for idx, row in equivalent_courses_df.iterrows():
        primary = row["Course"].strip().upper()
        equivalents = [x.strip().upper() for x in str(row["Equivalent"]).split(",")]
        for eq in equivalents:
            mapping[eq] = primary
    return mapping

def process_progress_report(
    df: pd.DataFrame,
    target_courses: dict,
    intensive_courses: dict,
    target_rules: dict,
    intensive_rules: dict,
    per_student_assignments: dict = None,
    equivalent_courses_mapping: dict = None
):
    """
    df: the raw long‐format progress data
    target_courses: { course_code: credits, ... }
    intensive_courses: { course_code: credits, ... }
    target_rules:    { course_code: [ {Credits, PassingGrades, FromOrd, ToOrd}, ... ], ... }
    intensive_rules: { course_code: [ {Credits, PassingGrades, FromOrd, ToOrd}, ... ], ... }
    per_student_assignments: { student_id: { assign_type: course, ... }, ... }
    equivalent_courses_mapping: { alt_code: primary_code, ... }
    """

    if equivalent_courses_mapping is None:
        equivalent_courses_mapping = {}

    # 1) Map equivalents
    df["Mapped Course"] = df["Course"].apply(lambda x: equivalent_courses_mapping.get(x, x))

    # 2) Apply S.C.E./F.E.C. (or any assignment types)
    if per_student_assignments:
        allowed_types = get_allowed_assignment_types()
        def map_assignment(row):
            sid = str(row["ID"])
            course = row["Course"]
            mapped = row["Mapped Course"]
            if sid in per_student_assignments:
                assigns = per_student_assignments[sid]
                for atype in allowed_types:
                    if assigns.get(atype) == course:
                        return atype
            return mapped

        df["Mapped Course"] = df.apply(map_assignment, axis=1)

    # 3) Compute term ordinal for each row
    df["TermOrd"] = df.apply(
        lambda r: semester_to_ordinal(r.get("Semester", ""), r.get("Year", "")),
        axis=1
    )

    # 4) Pre-format each row so CR is not lost and embed rule logic (with term matching)
    df["ProcessedValue"] = df.apply(
        lambda r: determine_course_value(
            r["Grade"],
            r["Mapped Course"],
            target_courses if r["Mapped Course"] in target_courses else
            intensive_courses if r["Mapped Course"] in intensive_courses else
            {},
            (target_rules.get(r["Mapped Course"], []) if r["Mapped Course"] in target_rules else
             intensive_rules.get(r["Mapped Course"], [])),
            r["TermOrd"]
        ),
        axis=1
    )

    # Capture the full student roster before splitting, so joins can retain
    # students even if they lack rows in a particular category (e.g., intensive).
    roster_df = df[["ID", "NAME"]].drop_duplicates()

    # 5) Split into required, intensive, extra
    extra_courses_df = df[
        (~df["Mapped Course"].isin(target_courses.keys())) &
        (~df["Mapped Course"].isin(intensive_courses.keys()))
    ]
    target_df = df[df["Mapped Course"].isin(target_courses.keys())]
    intensive_df = df[df["Mapped Course"].isin(intensive_courses.keys())]

    # 6) Pivot on ProcessedValue
    pivot_df = target_df.pivot_table(
        index=["ID", "NAME"],
        columns="Mapped Course",
        values="ProcessedValue",
        aggfunc=lambda vals: ", ".join(vals)
    ).reset_index()

    intensive_pivot_df = intensive_df.pivot_table(
        index=["ID", "NAME"],
        columns="Mapped Course",
        values="ProcessedValue",
        aggfunc=lambda vals: ", ".join(vals)
    ).reset_index()

    # Ensure the required and intensive tables retain all students from the original
    # roster, even if they lack rows in a particular category.
    pivot_df = roster_df.merge(pivot_df, on=["ID", "NAME"], how="left")
    intensive_pivot_df = roster_df.merge(intensive_pivot_df, on=["ID", "NAME"], how="left")

    # 7) Fill missing columns with "NR"
    for course in target_courses:
        if course not in pivot_df.columns:
            pivot_df[course] = "NR"
        else:
            pivot_df[course] = pivot_df[course].fillna("NR")

    for course in intensive_courses:
        if course not in intensive_pivot_df.columns:
            intensive_pivot_df[course] = "NR"
        else:
            intensive_pivot_df[course] = intensive_pivot_df[course].fillna("NR")

    result_df = pivot_df[["ID", "NAME"] + list(target_courses.keys())]
    intensive_result_df = intensive_pivot_df[["ID", "NAME"] + list(intensive_courses.keys())]

    # 7) Remove assigned courses from extras
    if per_student_assignments:
        assigned = [
            (sid, crs)
            for sid, assigns in per_student_assignments.items()
            for crs in assigns.values()
        ]
        extra_courses_df = extra_courses_df[
            ~extra_courses_df.apply(
                lambda row: (str(row["ID"]), row["Course"]) in assigned, axis=1
            )
        ]

    extra_courses_list = sorted(extra_courses_df["Course"].unique())
    return result_df, intensive_result_df, extra_courses_df, extra_courses_list

def semester_to_ordinal(semester: str, year) -> float:
    """
    Converts semester and year to an ordinal for comparison.
    Uses: year * 3 + {FALL: 0, SPRING: 1, SUMMER: 2}
    """
    try:
        yr = int(year)
        sem = str(semester).strip().upper()
        sem_map = {"FALL": 0, "SPRING": 1, "SUMMER": 2}
        return yr * 3 + sem_map.get(sem, 0)
    except (ValueError, TypeError):
        return float('-inf')


def determine_course_value(grade: str, course: str, courses_dict: dict, rules_list: list, term_ord: float = None):
    """
    Processes a course grade, taking into account:
      - Numeric credits (for non‐zero‐credit courses)
      - PASS/FAIL for zero‐credit courses
      - A list of rule‐dicts specifying FromOrd ≤ course_term_ord ≤ ToOrd and PassingGrades for each term‐range

    rules_list: [
      {"Credits": int, "PassingGrades": "A+,A,A-", "FromOrd": 10, "ToOrd": 16},
      {"Credits": int, "PassingGrades": "B+,B,B-", "FromOrd": 17, "ToOrd": 99},
      ...
    ]
    term_ord: The ordinal of the semester when the student took the course
    """

    credits = 0
    passing = ""

    if rules_list:
        matched_rule = None
        if term_ord is not None:
            for rule in rules_list:
                from_ord = rule.get("FromOrd", float('-inf'))
                to_ord = rule.get("ToOrd", float('inf'))
                if from_ord <= term_ord <= to_ord:
                    matched_rule = rule
                    break
        if matched_rule is None:
            matched_rule = rules_list[0]
        credits = matched_rule["Credits"]
        passing = matched_rule["PassingGrades"]

    if pd.isna(grade):
        return "NR"
    elif grade == "":
        return f"CR | {credits}" if credits > 0 else "CR | PASS"
    else:
        tokens = [g.strip().upper() for g in grade.split(", ") if g.strip()]
        all_toks = ", ".join(tokens)
        allowed = [x.strip().upper() for x in passing.split(",")] if passing else []
        passed = any(g in allowed for g in tokens)

        if credits > 0:
            return f"{all_toks} | {credits}" if passed else f"{all_toks} | 0"
        else:
            return f"{all_toks} | PASS" if passed else f"{all_toks} | FAIL"

def calculate_credits(row: pd.Series, courses_dict: dict):
    """
    Calculates Completed, Registered, Remaining, Total Credits.
    - If any "CR" appears in a multi‐attemp cell → count as Registered (not Completed).
    - If any token has numeric > 0 or PASS → count as Completed.
    - Else → Remaining.
    """
    completed, registered, remaining = 0, 0, 0
    total = sum(info for info in courses_dict.values())

    for course, cred in courses_dict.items():
        val = row.get(course, "")
        if isinstance(val, str):
            entries = [e.strip() for e in val.split(",") if e.strip()]

            # 1) Registered (CR precedence)
            if any(e.upper().startswith("CR") for e in entries):
                registered += cred
                continue

            # 2) Completed if any numeric > 0 or PASS
            passed = False
            for e in entries:
                parts = [p.strip() for p in e.split("|")]
                if len(parts) == 2:
                    tok, num = parts
                    try:
                        if int(num) > 0:
                            passed = True
                            break
                    except ValueError:
                        if num.upper() == "PASS":
                            passed = True
                            break

            if passed:
                completed += cred
            else:
                remaining += cred
        else:
            remaining += cred

    return pd.Series(
        [completed, registered, remaining, total],
        index=["# of Credits Completed", "# Registered", "# Remaining", "Total Credits"]
    )

def save_report_with_formatting(displayed_df: pd.DataFrame, intensive_displayed_df: pd.DataFrame, timestamp: str):
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font, Alignment
    from config import cell_color, COMPLETION_COLOR_MAP

    output = io.BytesIO()
    workbook = Workbook()
    ws_req = workbook.active
    ws_req.title = "Required Courses"

    completed_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    current_fill   = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    incomplete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(displayed_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_req.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if value == "c":
                    cell.fill = completed_fill
                elif value == "nc" or value == "":
                    cell.fill = incomplete_fill
                else:
                    style = cell_color(str(value))
                    if COMPLETION_COLOR_MAP["c"] in style:
                        cell.fill = completed_fill
                    elif COMPLETION_COLOR_MAP["cr"] in style:
                        cell.fill = current_fill
                    else:
                        cell.fill = incomplete_fill

    ws_int = workbook.create_sheet(title="Intensive Courses")
    for r_idx, row in enumerate(dataframe_to_rows(intensive_displayed_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_int.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if value == "c":
                    cell.fill = completed_fill
                elif value == "nc" or value == "":
                    cell.fill = incomplete_fill
                else:
                    style = cell_color(str(value))
                    if COMPLETION_COLOR_MAP["c"] in style:
                        cell.fill = completed_fill
                    elif COMPLETION_COLOR_MAP["cr"] in style:
                        cell.fill = current_fill
                    else:
                        cell.fill = incomplete_fill

    workbook.save(output)
    output.seek(0)
    return output

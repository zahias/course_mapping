import pandas as pd
import streamlit as st

def read_progress_report(filepath):
    try:
        if filepath.lower().endswith('.xlsx') or filepath.lower().endswith('.xls'):
            xls = pd.ExcelFile(filepath)
            if 'Progress Report' in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name='Progress Report')
                required_columns = {'ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester'}
                if not required_columns.issubset(df.columns):
                    st.error(f"The following required columns are missing in the 'Progress Report' sheet: {required_columns - set(df.columns)}")
                    return None
                return df
            else:
                st.info("No 'Progress Report' sheet found. Attempting to process the file as a wide format.")
                # Read the first sheet by default and attempt transformation
                df = pd.read_excel(xls)
                df = transform_wide_format(df)
                if df is None:
                    st.error("Failed to transform the file in wide format. Please ensure the file matches the expected structure.")
                return df
        elif filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
            # If expected columns exist, return as is; otherwise, try wide format
            if {'Course', 'Grade', 'Year', 'Semester'}.issubset(df.columns):
                return df
            else:
                st.info("CSV file does not contain the expected columns. Attempting to process as wide format.")
                df = transform_wide_format(df)
                return df
        else:
            st.error("File format not recognized. Please upload an Excel or CSV file.")
            return None
    except Exception as e:
        st.error(f"An error occurred while reading the file: {e}")
        return None

def transform_wide_format(df):
    # Check for identifying columns in the wide format
    if 'STUDENT ID' not in df.columns or not any(col.startswith('COURSE') for col in df.columns):
        st.error("The provided file does not match the expected wide format (missing 'STUDENT ID' or COURSE columns).")
        return None

    course_cols = [c for c in df.columns if c.startswith('COURSE')]
    id_vars = [c for c in df.columns if c not in course_cols]

    df_melted = df.melt(id_vars=id_vars, var_name='Course_Column', value_name='CourseData')
    df_melted = df_melted[df_melted['CourseData'].notnull() & (df_melted['CourseData'] != '')]

    split_cols = df_melted['CourseData'].str.split('/', expand=True)
    if split_cols.shape[1] < 3:
        st.error("Could not parse course data. Expected format: COURSECODE/SEMESTER-YEAR/GRADE.")
        return None

    df_melted['Course'] = split_cols[0].str.strip().str.upper()
    df_melted['Semester_Year'] = split_cols[1].str.strip()
    df_melted['Grade'] = split_cols[2].str.strip().str.upper()

    sem_year_split = df_melted['Semester_Year'].str.split('-', expand=True)
    if sem_year_split.shape[1] < 2:
        st.error("Semester-Year format not recognized. Expected something like FALL-2016.")
        return None
    df_melted['Semester'] = sem_year_split[0].str.strip().str.title()
    df_melted['Year'] = sem_year_split[1].str.strip()

    # Rename columns for consistency
    df_melted = df_melted.rename(columns={
        'STUDENT ID': 'ID',
        'NAME': 'NAME'
    })

    required_columns = {'ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester'}
    if not required_columns.issubset(df_melted.columns):
        st.error(f"The transformed data is missing some required columns: {required_columns - set(df_melted.columns)}")
        return None

    df_final = df_melted[['ID', 'NAME', 'Course', 'Grade', 'Year', 'Semester']].drop_duplicates()
    return df_final

def read_equivalent_courses(equivalent_courses_df):
    # Create a mapping dictionary
    mapping = {}
    for idx, row in equivalent_courses_df.iterrows():
        primary_course = row['Course'].strip().upper()
        equivalents = [course.strip().upper() for course in str(row['Equivalent']).split(',')]
        for eq_course in equivalents:
            mapping[eq_course] = primary_course  # Map equivalent course to primary course
    return mapping

def process_progress_report(
    df,
    target_courses,
    intensive_courses,
    grading_system,
    per_student_assignments=None,
    equivalent_courses_mapping=None  # Add this parameter
):
    # Map courses using the equivalent courses mapping
    if equivalent_courses_mapping is None:
        equivalent_courses_mapping = {}

    # Apply the mapping to the 'Course' column
    df['Mapped Course'] = df['Course'].apply(
        lambda x: equivalent_courses_mapping.get(x, x)
    )

    # Apply per-student S.C.E. and F.E.C. assignments
    if per_student_assignments:
        def map_sce_fec(row):
            student_id = str(row['ID'])
            course = row['Course']
            mapped_course = row['Mapped Course']

            # Check if this course is selected as S.C.E. or F.E.C. for this student
            if student_id in per_student_assignments:
                assignments = per_student_assignments[student_id]
                if assignments.get('S.C.E.') == course:
                    return 'S.C.E'
                elif assignments.get('F.E.C.') == course:
                    return 'F.E.C'
            return mapped_course

        df['Mapped Course'] = df.apply(map_sce_fec, axis=1)

    # Identify extra courses (exclude target and intensive courses)
    extra_courses_df = df[
        (~df['Mapped Course'].isin(target_courses)) &
        (~df['Mapped Course'].isin(intensive_courses))
    ]

    # Filter df to only target courses
    target_df = df[df['Mapped Course'].isin(target_courses)]

    # Filter df to only intensive courses
    intensive_df = df[df['Mapped Course'].isin(intensive_courses)]

    # Create pivot tables for target courses and intensive courses
    # Required courses pivot table
    pivot_df = target_df.pivot_table(
        index=['ID', 'NAME'],
        columns='Mapped Course',
        values='Grade',
        aggfunc=lambda x: ', '.join(map(str, filter(pd.notna, x)))
    ).reset_index()

    # Intensive courses pivot table
    intensive_pivot_df = intensive_df.pivot_table(
        index=['ID', 'NAME'],
        columns='Mapped Course',
        values='Grade',
        aggfunc=lambda x: ', '.join(map(str, filter(pd.notna, x)))
    ).reset_index()

    # Now process each required course
    for course in target_courses:
        if course not in pivot_df.columns:
            pivot_df[course] = None

        pivot_df[course] = pivot_df[course].apply(
            lambda grade: determine_course_value(grade, course, target_courses, grading_system)
        )

    # Process each intensive course
    for course in intensive_courses:
        if course not in intensive_pivot_df.columns:
            intensive_pivot_df[course] = None

        intensive_pivot_df[course] = intensive_pivot_df[course].apply(
            lambda grade: determine_course_value(grade, course, intensive_courses, grading_system)
        )

    # Keep only ID, NAME, and the courses
    result_df = pivot_df[['ID', 'NAME'] + list(target_courses.keys())]
    intensive_result_df = intensive_pivot_df[['ID', 'NAME'] + list(intensive_courses.keys())]

    # Remove assigned courses from extra_courses_df
    if per_student_assignments:
        assigned_courses = []
        for student_id, assignments in per_student_assignments.items():
            for course in assignments.values():
                assigned_courses.append((student_id, course))
        extra_courses_df = extra_courses_df[
            ~extra_courses_df.apply(lambda row: (str(row['ID']), row['Course']) in assigned_courses, axis=1)
        ]

    # Get list of extra courses (excluding intensive courses)
    extra_courses_list = sorted(extra_courses_df['Course'].unique())

    return result_df, intensive_result_df, extra_courses_df, extra_courses_list

def determine_course_value(grade, course, courses_dict, grading_system):
    if pd.isna(grade):
        return 'NR'  # Not Registered
    elif grade == '':
        return f'CR | {courses_dict[course]}'  # Currently Registered
    else:
        grades = grade.split(', ')
        grades_cleaned = [g.strip() for g in grades if g.strip()]
        all_grades = ', '.join(grades_cleaned)
        counted_grades = [g for g in grades_cleaned if g in grading_system['Counted']]
        if not counted_grades:
            return f'{all_grades} | 0'
        else:
            counted_credits = courses_dict[course]
            return f'{all_grades} | {counted_credits}'

def calculate_credits(row, courses_dict, grading_system):
    """
    Calculates the number of credits completed, registered, remaining, and total.
    """
    completed, registered, remaining = 0, 0, 0
    total_credits = sum(courses_dict.values())

    for course in courses_dict:
        value = row.get(course, '')
        if isinstance(value, str):
            value_upper = value.upper()
            if value_upper.startswith('CR'):
                # Currently Registered
                registered += courses_dict[course]
            elif value_upper.startswith('NR'):
                # Not Registered
                remaining += courses_dict[course]
            else:
                # Extract grades and determine if completed
                parts = value.split('|')
                grades_part = parts[0].strip()
                grades_list = [g.strip() for g in grades_part.split(',') if g.strip()]
                if any(grade in grading_system['Counted'] for grade in grades_list):
                    completed += courses_dict[course]
                else:
                    remaining += courses_dict[course]
        else:
            remaining += courses_dict[course]

    return pd.Series([completed, registered, remaining, total_credits],
                     index=['# of Credits Completed', '# Registered', '# Remaining', 'Total Credits'])


def save_report_with_formatting(displayed_df, intensive_displayed_df, timestamp, grading_system):
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font, Alignment

    output = io.BytesIO()
    workbook = Workbook()
    ws_required = workbook.active
    ws_required.title = "Required Courses"

    # Define fill colors
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')

    # Write the displayed_df to Excel with formatting
    for r_idx, row in enumerate(dataframe_to_rows(displayed_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_required.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                # Header row formatting
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Apply color formatting based on cell value
                if value == 'c':
                    cell.fill = light_green_fill
                elif value == '':
                    cell.fill = pink_fill
                else:
                    if isinstance(value, str):
                        grades_list = [g.strip() for g in value.split(',') if g.strip()]
                        if any(grade in grading_system['Counted'] or 'CR' == grade.upper() for grade in grades_list):
                            cell.fill = light_green_fill
                        else:
                            cell.fill = pink_fill

    # Add the Intensive Courses sheet
    ws_intensive = workbook.create_sheet(title="Intensive Courses")
    for r_idx, row in enumerate(dataframe_to_rows(intensive_displayed_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_intensive.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                # Header row formatting
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Apply color formatting based on cell value
                if value == 'c':
                    cell.fill = light_green_fill
                elif value == '':
                    cell.fill = pink_fill
                else:
                    if isinstance(value, str):
                        grades_list = [g.strip() for g in value.split(',') if g.strip()]
                        if any(grade in grading_system['Counted'] or 'CR' == grade.upper() for grade in grades_list):
                            cell.fill = light_green_fill
                        else:
                            cell.fill = pink_fill

    # Save the workbook to the output stream
    workbook.save(output)
    output.seek(0)
    return output

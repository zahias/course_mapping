# reporting.py

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List

def apply_excel_formatting(
    output: BytesIO,
    student_name: str,
    student_id: int,
    credits_completed: int,
    standing: str,
    note: str,
    advised_credits: int,
    optional_credits: int,
) -> None:
    """
    Enhances the single-student advising report workbook contained in `output`.
    Adds a header block (student info), formats the header row, freezes panes,
    and color-codes the 'Action' column including the new 'Registered' state.
    """
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Insert student info block
    ws.insert_rows(1, amount=8)
    ws["A1"] = "Student Advising Report"
    ws["A3"] = "Name:"; ws["B3"] = student_name
    ws["A4"] = "ID:"; ws["B4"] = student_id
    ws["A5"] = "Credits Completed:"; ws["B5"] = credits_completed
    ws["A6"] = "Standing:"; ws["B6"] = standing
    ws["A7"] = "Advisor Note:"; ws["B7"] = note
    ws["A8"] = "Credits (Advised / Optional):"; ws["B8"] = f"{advised_credits} / {optional_credits}"

    title_font = Font(size=14, bold=True)
    ws["A1"].font = title_font

    # Header row formatting
    header_row = 10
    ws.freeze_panes = f"A{header_row+1}"
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Borders for data region
    thin = Side(style="thin", color="CCCCCC")
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(header_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Conditional fill for Action column
    # Find "Action" header
    action_col_idx = None
    for idx, cell in enumerate(ws[header_row], start=1):
        if str(cell.value).strip().lower() == "action":
            action_col_idx = idx
            break
    if action_col_idx:
        fills = {
            "Completed": PatternFill("solid", fgColor="C6E0B4"),
            "Advised": PatternFill("solid", fgColor="FFF2CC"),
            "Optional": PatternFill("solid", fgColor="FFE699"),
            "Eligible (not chosen)": PatternFill("solid", fgColor="E1F0FF"),
            "Eligible not chosen": PatternFill("solid", fgColor="E1F0FF"),
            "Not Eligible": PatternFill("solid", fgColor="F8CECC"),
            "Registered": PatternFill("solid", fgColor="BDD7EE"),
        }
        for r in range(header_row + 1, max_row + 1):
            val = str(ws.cell(row=r, column=action_col_idx).value or "")
            fill = fills.get(val)
            if fill:
                ws.cell(row=r, column=action_col_idx).fill = fill

    output.seek(0)
    wb.save(output)
    output.seek(0)

def add_summary_sheet(writer, full_report: pd.DataFrame, course_cols: List[str]) -> None:
    """
    Add a 'Summary' sheet to a cohort report.
    Counts per-status for each course among the provided course columns.
    Recognized codes: c, r, a, na, ne
    """
    summary_rows = []
    for c in course_cols:
        counts = full_report[c].value_counts(dropna=False).to_dict()
        summary_rows.append({
            "Course": c,
            "Completed (c)": int(counts.get("c", 0)),
            "Registered (r)": int(counts.get("r", 0)),
            "Advised (a)": int(counts.get("a", 0)),
            "Eligible not chosen (na)": int(counts.get("na", 0)),
            "Not Eligible (ne)": int(counts.get("ne", 0)),
        })
    pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Summary")

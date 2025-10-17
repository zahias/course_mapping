# 5_Advising.py

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

from utilities import (
    check_course_completed,
    check_course_registered,
    is_course_offered,
    build_requisites_str,
    check_eligibility,
    get_student_total_credits,
    get_student_standing,
    style_df,
)

st.title("Advising")

# --------- Guards ---------
if "courses_df" not in st.session_state or st.session_state.courses_df is None or st.session_state.courses_df.empty:
    st.error("Combined courses catalog not loaded. Build it in **Customize Courses**.")
    st.stop()

if "progress_df" not in st.session_state or st.session_state.progress_df is None or st.session_state.progress_df.empty:
    st.error("Progress report not loaded. Upload it on the main page.")
    st.stop()

if "advising_selections" not in st.session_state:
    st.session_state.advising_selections = {}

courses_df = st.session_state.courses_df
progress_df = st.session_state.progress_df

# --------- Select Student ---------
student_list = progress_df["ID"].astype(str) + " - " + progress_df["NAME"]
selected_student = st.selectbox("Select Student", student_list)
selected_id = selected_student.split(" - ")[0]
student_row = progress_df[progress_df["ID"] == int(selected_id)].iloc[0].to_dict()

if selected_id not in st.session_state.advising_selections:
    st.session_state.advising_selections[selected_id] = {"advised": [], "optional": [], "note": ""}

current = st.session_state.advising_selections[selected_id]
cur_advised = current.get("advised", [])
cur_optional = current.get("optional", [])
cur_note = current.get("note", "")

# --------- Credits & Standing ---------
total_credits = get_student_total_credits(student_row)
standing = get_student_standing(total_credits)

c1, c2, c3 = st.columns(3)
c1.metric("Student", student_row["NAME"])
c2.metric("Total Credits", total_credits)
c3.metric("Standing", standing)

# --------- Eligibility map ---------
eligibility = {}
justifications = {}
for code in courses_df["Course Code"]:
    status, reason = check_eligibility(student_row, code, cur_advised, courses_df)
    eligibility[code] = status
    justifications[code] = reason

# Eligible candidates for advising
eligible_list = [
    c for c in courses_df["Course Code"]
    if eligibility[c] == "Eligible" or c in cur_advised or c in cur_optional
]

# --------- Advising form ---------
with st.form("advising_form"):
    st.subheader("Selections")

    advised_options = [c for c in eligible_list if eligibility[c] == "Eligible"]
    advised_default = [c for c in cur_advised if c in advised_options]
    advised_sel = st.multiselect("Advised Courses", options=advised_options, default=advised_default)

    optional_options = [c for c in eligible_list if eligibility[c] == "Eligible" and c not in advised_sel]
    optional_default = [c for c in cur_optional if c in optional_options]
    optional_sel = st.multiselect("Optional Courses", options=optional_options, default=optional_default)

    note_input = st.text_area("Advisor Note (Optional)", value=cur_note)

    if st.form_submit_button("Save Advising"):
        st.session_state.advising_selections[selected_id]["advised"] = sorted(advised_sel)
        st.session_state.advising_selections[selected_id]["optional"] = sorted(optional_sel)
        st.session_state.advising_selections[selected_id]["note"] = note_input.strip()
        st.success("Advising selections saved.")
        st.rerun()

# --------- Course Eligibility Table ---------
rows = []
for code in courses_df["Course Code"]:
    info = courses_df.loc[courses_df["Course Code"] == code].iloc[0]
    offered = "Yes" if is_course_offered(courses_df, code) else "No"
    typ = info.get("Type", "Required")
    reqs = build_requisites_str(info)

    status = eligibility[code]
    just = justifications[code]

    # derive action from progress or advising
    if check_course_completed(student_row, code):
        action = "Completed"
        status = "Completed"
    elif check_course_registered(student_row, code):
        action = "Registered"
        status = "Registered"
    elif code in advised_sel:
        action = "Advised"
    elif code in optional_sel:
        action = "Optional"
    else:
        action = "Eligible (not chosen)" if status == "Eligible" else "Not Eligible"

    if status == "Eligible" and not just:
        just = "All requirements met."

    rows.append({
        "Course Code": code,
        "Type": typ,
        "Requisites": reqs,
        "Eligibility Status": status,
        "Justification": just,
        "Offered": offered,
        "Action": action
    })

display_df = pd.DataFrame(rows)
req_df = display_df[display_df["Type"] == "Required"].copy()
int_df = display_df[display_df["Type"] == "Intensive"].copy()

st.markdown("### Course Eligibility")
if not req_df.empty:
    st.markdown("**Required Courses**")
    st.dataframe(style_df(req_df), use_container_width=True)
if not int_df.empty:
    st.markdown("**Intensive Courses**")
    st.dataframe(style_df(int_df), use_container_width=True)

# --------- Download Student Advising Report ---------

def export_student_report(df_required: pd.DataFrame, df_intensive: pd.DataFrame) -> BytesIO:
    combined = pd.concat([df_required, df_intensive], ignore_index=True)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Advising Report")
        wb = writer.book
        ws = writer.sheets["Advising Report"]

        # Insert header info (7 rows)
        ws.insert_rows(1, 7)
        ws["A1"] = "Student Name:"; ws["B1"] = student_row["NAME"]
        ws["A2"] = "Student ID:";   ws["B2"] = student_row["ID"]
        ws["A3"] = "# of Credits Completed:"; ws["B3"] = student_row.get("# of Credits Completed", 0)
        ws["A4"] = "# Registered:"; ws["B4"] = student_row.get("# Registered", 0)
        ws["A5"] = "Total Credits:"; ws["B5"] = total_credits
        ws["A6"] = "Standing:"; ws["B6"] = standing
        ws["A7"] = "Note:";     ws["B7"] = st.session_state.advising_selections[selected_id].get("note", "")

        # Bold table header
        header_row = 8
        for cell in ws[header_row]:
            cell.font = Font(bold=True)

        # Color rows based on Action
        # Find 'Action' column
        action_col_idx = None
        for idx, c in enumerate(ws[header_row], start=1):
            if c.value == "Action":
                action_col_idx = idx
                break

        fills = {
            "Completed": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
            "Registered": PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid"),
            "Advised": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "Optional": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
            "Eligible (not chosen)": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
            "Not Eligible": PatternFill(start_color="F08080", end_color="F08080", fill_type="solid"),
        }

        if action_col_idx:
            for r in range(header_row+1, ws.max_row+1):
                val = ws.cell(row=r, column=action_col_idx).value or ""
                # find matching key
                chosen = None
                for k, fill in fills.items():
                    if k in str(val):
                        chosen = fill
                        break
                if chosen:
                    for c in range(1, ws.max_column+1):
                        ws.cell(row=r, column=c).fill = chosen

        # Auto-adjust columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))

    output.seek(0)
    return output

st.markdown("### Export")
if st.button("Download Student Advising Report (Excel)"):
    buf = export_student_report(req_df, int_df)
    st.download_button(
        "Click to Download",
        data=buf.getvalue(),
        file_name=f"{student_row['NAME'].replace(' ','_')}_Advising_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
